"""
Martin Capital Partners -- Dividend Streak & CCC Data Lookup
data/dividend_streaks.py

Reads comprehensive dividend data from the David Fish / IREIT CCC spreadsheet
(Fish_*.xlsx). Ryan downloads this monthly from ireitinvestor.com.

Data extracted:
  "All CCC" sheet:
    - Consecutive years of increases (col 4)
    - DGR 1/3/5/10-year (cols 18-21) -- replaces unreliable yfinance growth rates
    - EPS Payout Ratio (col 25) -- replaces unreliable yfinance payout ratio
    - Chowder Rule (col 41) -- yield + 5Y DGR
    - Streak began year (col 56)
    - Recessions survived (col 57)

  "Historical" sheet:
    - Annual dividend per share going back to 1999 (cols 2-30, years in row 5)
    - Year-over-year pct increases (cols 32+)
"""

import os
import glob
import streamlit as st

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_THIS_DIR)

FISH_PATTERNS = [
    os.path.join(_THIS_DIR, "Fish_*.xlsx"),
    os.path.join(_THIS_DIR, "CCC_*.xlsx"),
    os.path.join(_THIS_DIR, "fish_*.xlsx"),
    os.path.join(_PROJECT_ROOT, "data", "Fish_*.xlsx"),
    os.path.join(_PROJECT_ROOT, "data", "CCC_*.xlsx"),
    "data/Fish_*.xlsx",
    "data/CCC_*.xlsx",
]


def _find_newest_fish():
    candidates = []
    for pattern in FISH_PATTERNS:
        candidates.extend(glob.glob(pattern))
    if not candidates:
        return None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _classify_tier(years):
    if years >= 50:
        return "King"
    elif years >= 25:
        return "Champion"
    elif years >= 10:
        return "Contender"
    elif years >= 5:
        return "Challenger"
    return "--"


def _sf(val):
    """Safe float conversion."""
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


@st.cache_data(ttl=300, show_spinner=False)
def _load_fish_data():
    """
    Load all CCC data from the Fish spreadsheet.
    Returns dict: {
        "streaks": {ticker: (years, tier)},
        "metrics": {ticker: {dgr_1y, dgr_3y, dgr_5y, dgr_10y, payout_ratio,
                             chowder, streak_began, recessions_survived}},
        "history": {ticker: {year: dividend_per_share}},
    }
    """
    result = {"streaks": {}, "metrics": {}, "history": {}}

    fish_path = _find_newest_fish()
    if not fish_path:
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(fish_path, read_only=True, data_only=True)
    except Exception:
        return result

    # ── Parse "All CCC" sheet ──────────────────────────────────────────
    if "All CCC" in wb.sheetnames:
        ws = wb["All CCC"]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (row with "Symbol" in it)
        header_row = 5  # default for known Fish format
        for i, row in enumerate(rows[:10]):
            if row and len(row) > 4:
                row_strs = [str(v).strip().lower() if v else "" for v in row[:10]]
                if "symbol" in row_strs:
                    header_row = i
                    break

        for row in rows[header_row + 1:]:
            if not row or len(row) < 5:
                continue
            symbol = str(row[1]).strip() if row[1] else ""
            if not symbol:
                continue

            years = int(_sf(row[4])) if row[4] else 0
            if years <= 0:
                continue

            sym = symbol.upper()
            result["streaks"][sym] = (years, _classify_tier(years))

            result["metrics"][sym] = {
                "dgr_1y":       round(_sf(row[18]), 1) if len(row) > 18 else 0,
                "dgr_3y":       round(_sf(row[19]), 1) if len(row) > 19 else 0,
                "dgr_5y":       round(_sf(row[20]), 1) if len(row) > 20 else 0,
                "dgr_10y":      round(_sf(row[21]), 1) if len(row) > 21 else 0,
                "payout_ratio": round(_sf(row[25]), 1) if len(row) > 25 else 0,
                "chowder":      round(_sf(row[41]), 1) if len(row) > 41 else 0,
                "streak_began": row[56] if len(row) > 56 and row[56] else None,
                "recessions":   int(_sf(row[57])) if len(row) > 57 else 0,
                "div_amount":   round(_sf(row[12]), 4) if len(row) > 12 else 0,
                "qtly_div":     round(_sf(row[10]), 4) if len(row) > 10 else 0,
            }

    # ── Parse "Historical" sheet ───────────────────────────────────────
    if "Historical" in wb.sheetnames:
        ws_hist = wb["Historical"]
        hist_rows = list(ws_hist.iter_rows(values_only=True))

        # Row 5 has year headers: col 0=Name, col 1=Symbol, cols 2-30 = years
        if len(hist_rows) > 5:
            year_headers = hist_rows[5]

            # Build year map: col_index -> year (int)
            year_map = {}
            for j in range(2, min(31, len(year_headers))):
                yr = year_headers[j]
                if yr is not None:
                    yr_str = str(yr).strip()
                    # Skip non-year columns like "#" or "RegDivs"
                    try:
                        yr_int = int(float(yr_str))
                        if 1990 <= yr_int <= 2030:
                            year_map[j] = yr_int
                    except (ValueError, TypeError):
                        pass

            # Parse data rows (start after header row 5)
            for row in hist_rows[6:]:
                if not row or len(row) < 3:
                    continue
                symbol = str(row[1]).strip() if row[1] else ""
                if not symbol:
                    continue

                sym = symbol.upper()
                div_history = {}
                for col_idx, year in year_map.items():
                    if col_idx < len(row) and row[col_idx] is not None:
                        val = _sf(row[col_idx])
                        if val > 0:
                            div_history[year] = round(val, 4)

                if div_history:
                    result["history"][sym] = div_history

    wb.close()
    return result


# ── Public API ─────────────────────────────────────────────────────────

def get_streak(ticker):
    """Get (years, tier) for a ticker. Returns (0, "--") if not found."""
    data = _load_fish_data()
    return data["streaks"].get(ticker.upper(), (0, "--"))


def get_streak_years(ticker):
    """Get just the number of consecutive years."""
    years, _ = get_streak(ticker)
    return years


def get_streak_tier(ticker):
    """Get just the tier label."""
    _, tier = get_streak(ticker)
    return tier


def get_fish_metrics(ticker):
    """
    Get all CCC metrics for a ticker.
    Returns dict with keys: dgr_1y, dgr_3y, dgr_5y, dgr_10y,
    payout_ratio, chowder, streak_began, recessions
    Returns empty dict if not found.
    """
    data = _load_fish_data()
    return data["metrics"].get(ticker.upper(), {})


def get_dividend_history(ticker):
    """
    Get annual dividend-per-share history from the Historical sheet.
    Returns dict: {year: amount} sorted by year, e.g. {1999: 0.545, 2000: 0.62, ...}
    Returns empty dict if not found.
    """
    data = _load_fish_data()
    hist = data["history"].get(ticker.upper(), {})
    return dict(sorted(hist.items()))


def get_all_fish_data(ticker):
    """
    Get everything for a ticker: streak, metrics, and history.
    Returns dict with keys: years, tier, metrics (dict), history (dict)
    """
    data = _load_fish_data()
    sym = ticker.upper()
    years, tier = data["streaks"].get(sym, (0, "--"))
    return {
        "years": years,
        "tier": tier,
        "metrics": data["metrics"].get(sym, {}),
        "history": data["history"].get(sym, {}),
    }


def get_all_streaks_for_tickers(tickers):
    """Get streak data for a list of tickers."""
    data = _load_fish_data()
    return {
        t: {
            "years": data["streaks"].get(t.upper(), (0, "--"))[0],
            "tier": data["streaks"].get(t.upper(), (0, "--"))[1],
        }
        for t in tickers
    }