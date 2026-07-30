"""
Martin Capital Partners — Warbook XLSX Export
data/warbook_export.py

Sprint 23D — Generates xlsx archives of the four warbook tabs (Strategy
Overview, QDG Characteristics, Risk Correlation, Attribution) for a single
strategy. Output is intended to match the printed warbook spreadsheet
format closely enough to run a one-quarter side-by-side comparison before
retiring the printed warbooks.

Layout per sheet (mirrors the printed warbook):
    Row 1     — Logo (top-left, A1) + Strategy name (bold, ~20pt)
    Row 2     — Tab subtitle (italic)
    Row 3     — Date (M/D/YYYY)
    Row 4     — Spacer
    Row 5     — Column headers (rotated 90° vertical, like the printed page)
    Row 6+    — Data rows, sorted per the on-screen tab convention

Print setup:
    Landscape, fit-to-width = 1 page, fit-to-height = 0 (multi-page tall),
    logo + title rows repeat on each printed page.

Missing data:
    Rendered as the literal em dash "—" (matches the on-screen and printed
    convention). Numeric cells use number_format; em dash cells stay text.

Public API:
    build_strategy_xlsx(...)         — returns BytesIO of all 4 tabs in one workbook
    build_single_tab_xlsx(...)       — returns BytesIO of one tab's workbook

Both functions take pre-fetched data dicts so the caller (warbook_tab.py)
can share its existing fetch results — no duplicated yfinance/Notion calls.
"""

from __future__ import annotations

import io
from datetime import date as _date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.drawing.image import Image as XLImage
    _XL_IMAGE_AVAILABLE = True
except ImportError:
    _XL_IMAGE_AVAILABLE = False

from utils.config import normalize_sector


# ── Module constants ──────────────────────────────────────────────────────

# Em dash — single source of truth so the on-screen and exported formats
# stay in lockstep. Matches the renderer's na_rep="—".
EM_DASH = "—"

# Default logo location. Caller can override via the logo_path parameter.
# Resolved relative to the project root (parent of `data/`).
_DEFAULT_LOGO = Path(__file__).resolve().parent.parent / "assets" / "M__Vector_.png"

# MCP brand colors — copied here rather than imported from utils.config so
# the export module has no Streamlit-dependent imports. Kept in sync with
# BRAND in utils/config.py manually.
_BRAND_GREEN  = "569542"
_BRAND_BLUE   = "07415A"
_BRAND_GOLD   = "C9A84C"
_BRAND_BLACK  = "0C1117"
_BRAND_RED    = "C45454"
_HEADER_FILL  = "F2F2F2"   # Light gray header band
_BORDER_GRAY  = "BFBFBF"

# Strategies that get a warbook (DCP excluded — mirrors warbook_tab.py)
WARBOOK_STRATEGIES = {"DAC", "OR", "QDVD", "SMID"}

# Strategy display names — used in the printed title row
STRATEGY_DISPLAY_NAMES = {
    "QDVD": "Quality Dividend Strategy",
    "DAC":  "Quality All-Cap Dividend Strategy",
    "SMID": "Quality SMID Dividend Strategy",
    "OR":   "Oregon Dividend Strategy",
}

# Tab subtitles — match the printed warbook header text
TAB_SUBTITLES = {
    "overview":     "Strategy Holdings Overview",
    "qdg":          "QDG Characteristics",
    "risk":         "Risk Correlation",
    "attribution":  "Supplemental Attribution & Quality Characteristics",
    "cld_notes":    "CLD Thesis Notes",
}

# Super sector mapping — duplicated from warbook_tab.py to keep this module
# free of Streamlit imports. Keep in sync.
_SUPER_SECTOR_BY_SECTOR = {
    "Materials":               "Cyclical",
    "Basic Materials":         "Cyclical",
    "Consumer Discretionary":  "Cyclical",
    "Consumer Cyclical":       "Cyclical",
    "Financials":              "Cyclical",
    "Financial Services":      "Cyclical",
    "Real Estate":             "Cyclical",
    "Communication Services":  "Sensitive",
    "Energy":                  "Sensitive",
    "Industrials":             "Sensitive",
    "Technology":              "Sensitive",
    "Consumer Staples":        "Defensive",
    "Consumer Defensive":      "Defensive",
    "Healthcare":              "Defensive",
    "Utilities":               "Defensive",
}


# ── Helpers ───────────────────────────────────────────────────────────────

def _is_num(v):
    """Mirror of warbook_tab._is_num — non-NaN numeric check."""
    if v is None:
        return False
    if not isinstance(v, (int, float)):
        return False
    return v == v


def _fmt_date_short(dt):
    """Format datetime as M/YY (cross-platform — avoids %-m / %#m)."""
    if dt is None:
        return ""
    try:
        return f"{dt.month}/{str(dt.year)[-2:]}"
    except (AttributeError, TypeError):
        return ""


def _fmt_date_md(dt):
    """Format datetime as M/D/YY (cross-platform)."""
    if dt is None:
        return ""
    try:
        return f"{dt.month}/{dt.day}/{str(dt.year)[-2:]}"
    except (AttributeError, TypeError):
        return ""


def _coerce_open_date(raw):
    """Coerce a Tamarac open_date (datetime, ISO string, or M/D/YYYY) to M/YY."""
    if not raw:
        return ""
    if isinstance(raw, datetime):
        return _fmt_date_short(raw)
    try:
        parsed = pd.to_datetime(raw, errors="coerce")
        if pd.notna(parsed):
            return _fmt_date_short(parsed)
        return str(raw)
    except Exception:
        return str(raw)


def _coerce_date_eval(raw):
    """Coerce a Notion ISO date string to M/D/YY."""
    if not raw:
        return ""
    try:
        d = datetime.strptime(raw, "%Y-%m-%d")
        return _fmt_date_md(d)
    except Exception:
        return str(raw)


# ── Row builders ──────────────────────────────────────────────────────────
#
# These mirror the rows = [] blocks in warbook_tab.py exactly. If a column
# changes there, mirror the change here (or — preferred — extract these into
# a shared helper module in a future sprint and have both files import it).
#
# Each builder returns a list of dicts (one per holding), pre-sorted to match
# the on-screen rendering order.

def _build_overview_rows(tam_df, price_data, notion_data, div_data):
    """Replicates _render_strategy_overview's row construction."""
    rows = []
    for _, h in tam_df.iterrows():
        sym = str(h["symbol"]).strip().upper()
        mkt = price_data.get(sym, {})
        nm = notion_data.get(sym, {})
        dd = div_data.get(sym, {})

        price = mkt.get("price") or 0
        unit_cost = h.get("unit_cost") or 0

        delta_from_cost = None
        if unit_cost and price:
            delta_from_cost = round((price - unit_cost) / unit_cost * 100, 2)

        mcp_target = nm.get("mcp_target")
        pct_to_target = None
        if mcp_target and price:
            pct_to_target = round((mcp_target - price) / price * 100, 1)

        growth_5y = dd.get("div_growth_5y")
        baseline = nm.get("div_baseline")

        exceeds = None
        if isinstance(growth_5y, (int, float)) and isinstance(baseline, (int, float)):
            exceeds = "Yes" if growth_5y >= baseline else "No"

        rows.append({
            "Company":      h["description"],
            "Symbol":       sym,
            "Weight":       round(h["weight_pct"], 2),
            "Sector":       normalize_sector(mkt.get("sector", "")),
            "Yield":        mkt.get("dividend_yield") or 0,
            "Open Date":    _coerce_open_date(h.get("open_date", "")),
            "Cost Basis":   round(unit_cost, 2) if unit_cost else None,
            "Close":        round(price, 2) if price else None,
            "Δ from Cost":  delta_from_cost,
            "CLD":          nm.get("cld"),
            "CLD Source":   nm.get("cld_source") or "",
            "Style":        nm.get("style_bucket") or "",
            "3yr Tgt":      mcp_target,
            "% To Tgt":     pct_to_target,
            "Baseline":     baseline,
            "5yr DG":       growth_5y,
            "DG ≥ Base":    exceeds or EM_DASH,
            "Date Eval":    _coerce_date_eval(nm.get("date_evaluated") or ""),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    numeric_cols = [
        "Weight", "Yield", "Cost Basis", "Close", "Δ from Cost",
        "CLD", "3yr Tgt", "% To Tgt", "Baseline", "5yr DG",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df.sort_values("Weight", ascending=False).reset_index(drop=True)


def _build_qdg_rows(tam_df, price_data, notion_data, warbook_data, fish_data, fish_history):
    """Replicates _render_qdg_characteristics's row construction."""
    rows = []
    for _, h in tam_df.iterrows():
        sym = str(h["symbol"]).strip().upper()
        mkt = price_data.get(sym, {})
        nm = notion_data.get(sym, {})
        wm = warbook_data.get(sym, {})
        fm = fish_data.get(sym, {})
        fh = fish_history.get(sym, {})

        price = mkt.get("price") or 0
        qty = h.get("quantity") or 0
        tam_value = h.get("value") or 0
        value = tam_value if tam_value else (qty * price if qty and price else 0)

        mkt_cap_raw = mkt.get("market_cap") or 0
        mkt_cap_bln = round(mkt_cap_raw / 1e9, 1) if mkt_cap_raw else None

        paid_since = None
        if fh:
            years_with_div = sorted(y for y, v in fh.items() if v and v > 0)
            if years_with_div:
                paid_since = years_with_div[0]

        # Sprint 24-6: Raised Since now from Notion (was Fish streak_began).
        raised_since_raw = nm.get("raised_since")
        raised_since = None
        if raised_since_raw is not None:
            try:
                raised_since = int(float(str(raised_since_raw)))
            except (ValueError, TypeError):
                pass

        last_bump = None
        if fh:
            current_year = _date.today().year
            valid_years = sorted(y for y, v in fh.items() if v and v > 0)
            past_years = [y for y in valid_years if y < current_year]
            if len(past_years) >= 2:
                latest_y = past_years[-1]
                prior_y = past_years[-2]
                latest_v = fh.get(latest_y)
                prior_v = fh.get(prior_y)
                if latest_v and prior_v and prior_v > 0:
                    last_bump = round((latest_v / prior_v - 1) * 100, 1)

        # Sprint 24-6: DG Notion-first, Fish-fallback (mirrors warbook_tab).
        def _dgr_with_fallback(notion_v, fish_v):
            if notion_v is not None:
                return notion_v
            if fish_v == 0:
                return None
            return fish_v

        dgr_1y = _dgr_with_fallback(nm.get("dgr_1y"), fm.get("dgr_1y"))
        dgr_3y = _dgr_with_fallback(nm.get("dgr_3y"), fm.get("dgr_3y"))
        dgr_5y = _dgr_with_fallback(nm.get("dgr_5y"), fm.get("dgr_5y"))

        payout = fm.get("payout_ratio")
        if payout == 0:
            payout = None

        rows.append({
            "Symbol":        sym,
            "Shares":        qty,
            "Value":         value,
            "Company":       h["description"],
            "Yield":         mkt.get("dividend_yield") or 0,
            "Mkt Cap $Bln":  mkt_cap_bln,
            "Sector":        normalize_sector(mkt.get("sector", "")),
            "ROE %":         wm.get("roe_ttm"),
            "LT D/Cap %":    wm.get("lt_debt_to_capital"),
            "Qual (S&P)":    nm.get("sp_quality") or "",
            "Paid Since":    paid_since,
            "Raised Since":  raised_since,
            "Timing":        nm.get("timing_of_raise") or "",
            "Freq":          wm.get("dividend_frequency") or "",
            "Payout %":      payout,
            "Last Bump %":   last_bump,
            "1Y DG %":       dgr_1y,
            "3Y DG %":       dgr_3y,
            "5Y DG %":       dgr_5y,
            "FCF Yld %":     wm.get("fcf_yield"),
            "Weight":        round(h["weight_pct"], 2),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    numeric_cols = [
        "Shares", "Value", "Yield", "Mkt Cap $Bln", "ROE %", "LT D/Cap %",
        "Paid Since", "Raised Since", "Payout %", "Last Bump %",
        "1Y DG %", "3Y DG %", "5Y DG %", "FCF Yld %", "Weight",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df.sort_values("Weight", ascending=False).reset_index(drop=True)


def _build_risk_rows(tam_df, price_data, notion_data, warbook_data):
    """Replicates _render_risk_correlation's row construction."""
    rows = []
    for _, h in tam_df.iterrows():
        sym = str(h["symbol"]).strip().upper()
        mkt = price_data.get(sym, {})
        nm = notion_data.get(sym, {})
        wm = warbook_data.get(sym, {})

        mkt_cap_raw = mkt.get("market_cap") or 0
        mkt_cap_bln = round(mkt_cap_raw / 1e9, 1) if mkt_cap_raw else None

        sector_normalized = normalize_sector(mkt.get("sector", ""))
        super_sector = (
            _SUPER_SECTOR_BY_SECTOR.get(sector_normalized)
            or wm.get("super_sector")
            or ""
        )

        rows.append({
            "Symbol":         sym,
            "Close":          mkt.get("price") or 0,
            "Mkt Cap $Bln":   mkt_cap_bln,
            "Super Sector":   super_sector,
            "Sector":         sector_normalized,
            "Sub-Industry":   wm.get("sub_industry") or "",
            "Credit (S&P)":   nm.get("sp_credit") or "",
            "Debt Cov":       wm.get("debt_coverage_ratio"),
            "LT D/Cap %":     wm.get("lt_debt_to_capital"),
            "Beta":           mkt.get("beta") if mkt.get("beta") else None,
            "Style":          nm.get("mstar_style") or "",
            "Mstar Gr":       nm.get("mstar_growth") or "",
            "Mstar Pf":       nm.get("mstar_profitability") or "",
            "Mstar FH":       nm.get("mstar_fin_health") or "",
            "Country":        wm.get("country") or "",
            "Weight":         round(h["weight_pct"], 2),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    numeric_cols = ["Close", "Mkt Cap $Bln", "Debt Cov", "LT D/Cap %", "Beta", "Weight"]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df.sort_values(
        by=["Super Sector", "Sector", "Symbol"],
        ascending=[True, True, True],
        na_position="last",
    ).reset_index(drop=True)


def _build_attribution_rows(tam_df, price_data, warbook_data):
    """Replicates _render_attribution's row construction."""
    rows = []
    for _, h in tam_df.iterrows():
        sym = str(h["symbol"]).strip().upper()
        mkt = price_data.get(sym, {})
        wm = warbook_data.get(sym, {})

        price = mkt.get("price") or 0
        w52_high = mkt.get("52w_high") or 0
        from_52w_high = None
        if price and w52_high:
            from_52w_high = round((price - w52_high) / w52_high * 100, 1)

        qty = h.get("quantity") or 0
        tam_value = h.get("value") or 0
        value = tam_value if tam_value else (qty * price if qty and price else 0)

        rows.append({
            "Symbol":          sym,
            "Shares":          qty,
            "Value":           value,
            "Company":         h["description"],
            "YTD TR":          wm.get("tr_ytd"),
            "3M TR":           wm.get("tr_3m"),
            "1Y TR":           wm.get("tr_1y"),
            "MTD TR":          wm.get("tr_mtd"),
            "QTD TR":          wm.get("tr_qtd"),
            "QTD vs SPX":      wm.get("tr_qtd_vs_spx"),
            "YTD vs SPX":      wm.get("tr_ytd_vs_spx"),
            "% From 52W Hi":   from_52w_high,
            "% Net Debt/Cap":  wm.get("net_debt_to_capital"),
            "ROE 5Y Avg":      wm.get("roe_5y_avg"),
            "EPS Cov":         wm.get("eps_div_coverage"),
            "CF Cov":          wm.get("cf_div_coverage"),
            "FCF Cov":         wm.get("fcf_div_coverage"),
            "FWD P/E":         mkt.get("forward_pe") or wm.get("forward_pe"),
            "CF/EV Yield":     wm.get("cash_flow_ev_yield"),
            "Weight":          round(h["weight_pct"], 2),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    numeric_cols = [
        "Shares", "Value", "YTD TR", "3M TR", "1Y TR", "MTD TR", "QTD TR",
        "QTD vs SPX", "YTD vs SPX", "% From 52W Hi", "% Net Debt/Cap",
        "ROE 5Y Avg", "EPS Cov", "CF Cov", "FCF Cov", "FWD P/E",
        "CF/EV Yield", "Weight",
    ]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df.sort_values("YTD TR", ascending=False, na_position="last").reset_index(drop=True)


# ── Per-column formatting metadata ────────────────────────────────────────
#
# For each tab, define the Excel number_format and width for each column. The
# rendering loop below uses this to set cell number formats so totals look
# right when a printed warbook is held up next to the export.
#
# Number formats use Excel's syntax:
#   "0.0%"       — 12.3%
#   "+0.0%;-0.0%;0.0%"  — signed percent, zero stays zero
#   "$#,##0.00"  — $12,345.67
#   "0.0\"x\""   — 1.4x  (literal "x" suffix)
#   "@"          — text (forces no numeric coercion)

# Format codes
F_PCT_2     = "0.00%"
F_PCT_1     = "0.0%"
F_PCT_0     = "0%"
F_PCT_SIGN1 = "+0.0%;-0.0%;0.0%"
F_PCT_SIGN2 = "+0.00%;-0.00%;0.00%"
F_USD_0     = "$#,##0"
F_USD_2     = "$#,##0.00"
F_NUM_0     = "#,##0"
F_NUM_1     = "#,##0.0"
F_NUM_2     = "#,##0.00"
F_MULT_1    = "0.0\"x\""
F_TEXT      = "@"
F_INT       = "0"

# Alignment tokens — one of these per column, applied uniformly to data
# cells in that column. Keeps the export consistent (no length-based
# heuristics that produce inconsistent justification within a column).
A_LEFT   = "left"
A_CENTER = "center"
A_RIGHT  = "right"

# Column metadata: (label, key, width, number_format, alignment)
#
# Alignment convention (mirrors the printed warbook):
#   - Long descriptive text (Company, Sub-Industry, CLD Source) → left
#   - Short text codes (Symbol, Sector, ratings, dates, Yes/No) → center
#   - All numeric columns                                       → right

OVERVIEW_COLS = [
    ("Company",     "Company",     32, F_TEXT,      A_LEFT),
    ("Symbol",      "Symbol",      8,  F_TEXT,      A_CENTER),
    ("Weight",      "Weight",      8,  F_PCT_2,     A_RIGHT),
    ("Sector",      "Sector",      14, F_TEXT,      A_CENTER),
    ("Yield",       "Yield",       8,  F_PCT_2,     A_RIGHT),
    ("Open Date",   "Open Date",   10, F_TEXT,      A_CENTER),
    ("Cost Basis",  "Cost Basis",  12, F_USD_2,     A_RIGHT),
    ("Close",       "Close",       12, F_USD_2,     A_RIGHT),
    ("Δ from Cost", "Δ from Cost", 11, F_PCT_SIGN1, A_RIGHT),
    ("CLD",         "CLD",         8,  F_INT,       A_RIGHT),
    ("CLD Source",  "CLD Source",  14, F_TEXT,      A_LEFT),
    ("Style",       "Style",       8,  F_TEXT,      A_CENTER),
    ("3yr Tgt",     "3yr Tgt",     10, F_USD_0,     A_RIGHT),
    ("% To Tgt",    "% To Tgt",    10, F_PCT_SIGN1, A_RIGHT),
    ("Baseline",    "Baseline",    10, F_PCT_0,     A_RIGHT),
    ("5yr DG",      "5yr DG",      10, F_PCT_1,     A_RIGHT),
    ("DG ≥ Base",   "DG ≥ Base",   10, F_TEXT,      A_CENTER),
    ("Date Eval",   "Date Eval",   12, F_TEXT,      A_CENTER),
]

QDG_COLS = [
    ("Symbol",        "Symbol",       8,  F_TEXT,      A_CENTER),
    ("Shares",        "Shares",       10, F_NUM_0,     A_RIGHT),
    ("Value",         "Value",        14, F_USD_0,     A_RIGHT),
    ("Company",       "Company",      32, F_TEXT,      A_LEFT),
    ("Yield",         "Yield",        8,  F_PCT_2,     A_RIGHT),
    ("Mkt Cap $Bln",  "Mkt Cap $Bln", 11, F_NUM_1,     A_RIGHT),     # Stored already-in-billions; one decimal matches printed warbook
    ("Sector",        "Sector",       14, F_TEXT,      A_CENTER),
    ("ROE %",         "ROE %",        9,  F_PCT_1,     A_RIGHT),
    ("LT D/Cap %",    "LT D/Cap %",   10, F_PCT_1,     A_RIGHT),
    ("Qual (S&P)",    "Qual (S&P)",   10, F_TEXT,      A_CENTER),
    ("Paid Since",    "Paid Since",   10, F_INT,       A_CENTER),
    ("Raised Since",  "Raised Since", 11, F_INT,       A_CENTER),
    ("Timing",        "Timing",       8,  F_TEXT,      A_CENTER),
    ("Freq",          "Freq",         7,  F_TEXT,      A_CENTER),
    ("Payout %",      "Payout %",     10, F_PCT_1,     A_RIGHT),
    ("Last Bump %",   "Last Bump %",  11, F_PCT_SIGN1, A_RIGHT),
    ("1Y DG %",       "1Y DG %",      9,  F_PCT_SIGN1, A_RIGHT),
    ("3Y DG %",       "3Y DG %",      9,  F_PCT_SIGN1, A_RIGHT),
    ("5Y DG %",       "5Y DG %",      9,  F_PCT_SIGN1, A_RIGHT),
    ("FCF Yld %",     "FCF Yld %",    10, F_PCT_2,     A_RIGHT),
    ("Weight",        "Weight",       8,  F_PCT_2,     A_RIGHT),
]

RISK_COLS = [
    ("Symbol",         "Symbol",        8,  F_TEXT,    A_CENTER),
    ("Close",          "Close",         12, F_USD_2,   A_RIGHT),
    ("Mkt Cap $Bln",   "Mkt Cap $Bln",  11, F_NUM_1,   A_RIGHT),
    ("Super Sector",   "Super Sector",  12, F_TEXT,    A_CENTER),
    ("Sector",         "Sector",        14, F_TEXT,    A_CENTER),
    ("Sub-Industry",   "Sub-Industry",  22, F_TEXT,    A_LEFT),
    ("Credit (S&P)",   "Credit (S&P)",  10, F_TEXT,    A_CENTER),
    ("Debt Cov",       "Debt Cov",      10, F_MULT_1,  A_RIGHT),
    ("LT D/Cap %",     "LT D/Cap %",    10, F_PCT_1,   A_RIGHT),
    ("Beta",           "Beta",          7,  F_NUM_2,   A_RIGHT),
    ("Style",          "Style",         8,  F_TEXT,    A_CENTER),
    ("Mstar Gr",       "Mstar Gr",      9,  F_TEXT,    A_CENTER),
    ("Mstar Pf",       "Mstar Pf",      9,  F_TEXT,    A_CENTER),
    ("Mstar FH",       "Mstar FH",      9,  F_TEXT,    A_CENTER),
    ("Country",        "Country",       9,  F_TEXT,    A_CENTER),
    ("Weight",         "Weight",        8,  F_PCT_2,   A_RIGHT),
]

ATTRIBUTION_COLS = [
    ("Symbol",         "Symbol",         8,  F_TEXT,      A_CENTER),
    ("Shares",         "Shares",         10, F_NUM_0,     A_RIGHT),
    ("Value",          "Value",          14, F_USD_0,     A_RIGHT),
    ("Company",        "Company",        32, F_TEXT,      A_LEFT),
    ("YTD TR",         "YTD TR",         9,  F_PCT_SIGN1, A_RIGHT),
    ("3M TR",          "3M TR",          9,  F_PCT_SIGN1, A_RIGHT),
    ("1Y TR",          "1Y TR",          9,  F_PCT_SIGN1, A_RIGHT),
    ("MTD TR",         "MTD TR",         9,  F_PCT_SIGN1, A_RIGHT),
    ("QTD TR",         "QTD TR",         9,  F_PCT_SIGN1, A_RIGHT),
    ("QTD vs SPX",     "QTD vs SPX",     10, F_PCT_SIGN1, A_RIGHT),
    ("YTD vs SPX",     "YTD vs SPX",     10, F_PCT_SIGN1, A_RIGHT),
    ("% From 52W Hi",  "% From 52W Hi",  12, F_PCT_SIGN1, A_RIGHT),
    ("% Net Debt/Cap", "% Net Debt/Cap", 12, F_PCT_1,     A_RIGHT),
    ("ROE 5Y Avg",     "ROE 5Y Avg",     11, F_PCT_1,     A_RIGHT),
    ("EPS Cov",        "EPS Cov",        9,  F_MULT_1,    A_RIGHT),
    ("CF Cov",         "CF Cov",         9,  F_MULT_1,    A_RIGHT),
    ("FCF Cov",        "FCF Cov",        9,  F_MULT_1,    A_RIGHT),
    ("FWD P/E",        "FWD P/E",        9,  F_NUM_1,     A_RIGHT),
    ("CF/EV Yield",    "CF/EV Yield",    10, F_PCT_1,     A_RIGHT),
    ("Weight",         "Weight",         8,  F_PCT_2,     A_RIGHT),
]

# Columns whose negative values should render in red (matches the printed
# warbook's red parentheses convention). Defined per-tab.
RED_NEG_COLS = {
    "overview":     {"Δ from Cost", "% To Tgt"},
    "qdg":          set(),
    "risk":         set(),
    "attribution":  {"YTD TR", "3M TR", "1Y TR", "MTD TR", "QTD TR",
                     "QTD vs SPX", "YTD vs SPX", "% From 52W Hi"},
}

# Percent columns whose data is stored as already-multiplied numbers (e.g. 5.2
# meaning 5.2%) rather than 0.052. The warbook tabs use the human-readable
# convention everywhere, so all percent columns get divided by 100 when written
# to Excel so the F_PCT_* number formats render correctly. This set lists the
# label-keys whose RAW values need to be divided. (i.e. all percent columns.)
#
# Yield in price_data is also stored as a percent already (e.g. 3.01 not 0.0301).
#
# Mkt Cap $Bln uses F_USD_0 / F_NUM_1 — it's a number-of-billions, so no
# divide.
PCT_RAW_COLS = {
    "overview":     {"Weight", "Yield", "Δ from Cost", "% To Tgt", "Baseline", "5yr DG"},
    "qdg":          {"Yield", "ROE %", "LT D/Cap %", "Payout %", "Last Bump %",
                     "1Y DG %", "3Y DG %", "5Y DG %", "FCF Yld %", "Weight"},
    "risk":         {"LT D/Cap %", "Weight"},
    "attribution":  {"YTD TR", "3M TR", "1Y TR", "MTD TR", "QTD TR",
                     "QTD vs SPX", "YTD vs SPX", "% From 52W Hi",
                     "% Net Debt/Cap", "ROE 5Y Avg", "CF/EV Yield", "Weight"},
}


# ── Sheet writer ──────────────────────────────────────────────────────────

_THIN_SIDE = Side(style="thin", color=_BORDER_GRAY)
_HEADER_BORDER = Border(top=_THIN_SIDE, bottom=_THIN_SIDE, left=_THIN_SIDE, right=_THIN_SIDE)
_DATA_BORDER   = Border(bottom=Side(style="hair", color=_BORDER_GRAY))


def _write_tab(
    ws,
    tab_key,
    cols,
    df,
    strategy_code,
    strategy_display_name,
    as_of_date,
    logo_path=None,
):
    """
    Write one tab's worth of data to a worksheet.

    Layout:
      Row 1:  Logo (anchored A1, ~36px tall) + Strategy name in column B+
      Row 2:  Tab subtitle in column B+ (italic)
      Row 3:  As-of date in column B+
      Row 4:  blank spacer
      Row 5:  Headers (rotated 90° vertical)
      Row 6+: Data

    Title text is placed in column B onward so the logo (anchored to column A)
    has its own dedicated horizontal space and never overlaps text. Rows 1-3
    are made tall enough to give the logo vertical room.
    """
    HEADER_ROW = 5
    DATA_START_ROW = 6

    # ── Title block ────────────────────────────────────────────────────
    # Place title text starting at column B so the logo (anchored A1) has
    # its own column. Don't merge across columns — that interacts badly
    # with print fit-to-width on some Excel versions.
    ws["B1"] = f"Martin Capital Partners — {strategy_display_name}"
    ws["B1"].font = Font(name="Arial", size=16, bold=True, color=_BRAND_BLACK)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["B2"] = TAB_SUBTITLES.get(tab_key, "")
    ws["B2"].font = Font(name="Arial", size=11, italic=True, color="595959")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    date_str = _safe_strftime(as_of_date) or f"{as_of_date.month}/{as_of_date.day}/{as_of_date.year}"
    ws["B3"] = f"As of {date_str}"
    ws["B3"].font = Font(name="Arial", size=10, color="808080")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    # Make rows 1-3 tall enough that the logo (~50px = ~38pt) clears them
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    # Logo anchored at A1. Sized small enough to fit within column A's width
    # without bleeding into the title text in column B.
    if logo_path and _XL_IMAGE_AVAILABLE:
        try:
            p = Path(logo_path)
            if p.exists():
                img = XLImage(str(p))
                # Target ~50px tall — fits the 60pt total height of rows 1-3
                # (26+18+16 = 60pt ≈ 80px) with margin. The width is derived
                # from the original aspect ratio so the logo doesn't squash.
                target_h_px = 50
                if img.height and img.width:
                    aspect = img.width / img.height
                    img.height = target_h_px
                    img.width = int(target_h_px * aspect)
                img.anchor = "A1"
                ws.add_image(img)
        except Exception:
            # Logo failures should never break the export
            pass

    # ── Headers ────────────────────────────────────────────────────────
    header_font = Font(name="Arial", size=10, bold=True, color=_BRAND_BLACK)
    header_fill = PatternFill("solid", start_color=_HEADER_FILL, end_color=_HEADER_FILL)
    header_align = Alignment(
        horizontal="center", vertical="bottom",
        text_rotation=90, wrap_text=False,
    )

    for i, (label, _key, width, _fmt, _align) in enumerate(cols, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _HEADER_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    # Tall header row to give vertical labels room
    ws.row_dimensions[HEADER_ROW].height = 110

    # ── Data ───────────────────────────────────────────────────────────
    pct_raw = PCT_RAW_COLS.get(tab_key, set())
    red_neg = RED_NEG_COLS.get(tab_key, set())

    body_font_base = Font(name="Arial", size=10, color=_BRAND_BLACK)
    body_font_red  = Font(name="Arial", size=10, color=_BRAND_RED)

    # Pre-build alignment objects (one per align token) — reused across all
    # data cells so we don't allocate a new Alignment per cell.
    _align_objs = {
        A_LEFT:   Alignment(horizontal="left",   vertical="center"),
        A_CENTER: Alignment(horizontal="center", vertical="center"),
        A_RIGHT:  Alignment(horizontal="right",  vertical="center"),
    }

    for r_offset, (_, row) in enumerate(df.iterrows()):
        excel_row = DATA_START_ROW + r_offset
        for c_idx, (label, key, _w, fmt, align) in enumerate(cols, start=1):
            raw_val = row.get(key)
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.border = _DATA_BORDER
            # Per-column alignment — applies whether the cell ends up text,
            # numeric, or em dash. Keeps every value in a column in lockstep.
            cell.alignment = _align_objs[align]
            cell.font = body_font_base

            if not _is_num(raw_val):
                # Text or em dash — preserve the column's intended alignment
                if raw_val is None or (isinstance(raw_val, float) and raw_val != raw_val):
                    cell.value = EM_DASH
                else:
                    s = str(raw_val).strip()
                    cell.value = s if s else EM_DASH
            else:
                # Numeric — divide by 100 for percent columns whose source is
                # stored as already-multiplied (e.g. 5.2 meaning 5.2%).
                v = float(raw_val)
                if label in pct_raw:
                    v = v / 100.0
                cell.value = v
                cell.number_format = fmt
                # Red text for negative values in flagged columns
                if label in red_neg and v < 0:
                    cell.font = body_font_red

    # ── Print setup ────────────────────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True
    # Repeat title + header rows on every printed page
    ws.print_title_rows = f"1:{HEADER_ROW}"
    # Freeze panes — header + first column on screen
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=2)


def _safe_strftime(d):
    """
    Try a strftime that works on both Linux (%-m) and Windows (%#m) — fall
    back to attribute-based formatting if neither is available.
    """
    try:
        return d.strftime("%-m/%-d/%Y")
    except Exception:
        try:
            return d.strftime("%#m/%#d/%Y")
        except Exception:
            return None


# ── Public API ────────────────────────────────────────────────────────────

def _build_cld_notes_rows(tam_df, notion_data):
    """
    Symbol / Company / CLD Thesis for holdings that have commentary in the
    Notion "CLD Comments" column. Weight-sorted to match the on-screen
    thesis-notes panel. Holdings without commentary are omitted (blank rows
    add nothing to a printed research page).
    """
    rows = []
    for _, h in tam_df.iterrows():
        sym = str(h["symbol"]).strip().upper()
        cmt = (notion_data.get(sym, {}).get("cld_comments") or "").strip()
        if cmt:
            rows.append({
                "Weight":     h["weight_pct"],
                "Symbol":     sym,
                "Company":    h["description"],
                "CLD Thesis": cmt,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).sort_values("Weight", ascending=False)
    return df.drop(columns=["Weight"]).reset_index(drop=True)


def _write_cld_notes_tab(ws, df, strategy_code, strategy_display_name,
                         as_of_date, logo_path=None):
    """
    Dedicated writer for the thesis-notes sheet. Unlike the grid tabs
    (rotated headers, landscape, fit-to-width across 18 columns), thesis
    text wants reading width: portrait, three columns, wrapped text, and
    unset row heights so Excel auto-fits each thesis. Title block matches
    _write_tab so the printed page carries the same letterhead.
    """
    HEADER_ROW = 5
    DATA_START_ROW = 6

    ws["B1"] = f"Martin Capital Partners — {strategy_display_name}"
    ws["B1"].font = Font(name="Arial", size=16, bold=True, color=_BRAND_BLACK)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = TAB_SUBTITLES["cld_notes"]
    ws["B2"].font = Font(name="Arial", size=11, italic=True, color="595959")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    date_str = _safe_strftime(as_of_date) or f"{as_of_date.month}/{as_of_date.day}/{as_of_date.year}"
    ws["B3"] = f"As of {date_str}"
    ws["B3"].font = Font(name="Arial", size=10, color="808080")
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    if logo_path:
        try:
            img = XLImage(logo_path)
            if img.height and img.width:
                aspect = img.width / img.height
                target_h_px = 50
                img.height = target_h_px
                img.width = int(target_h_px * aspect)
            ws.add_image(img, "A1")
        except Exception:
            pass

    cols = [("Symbol", 9), ("Company", 30), ("CLD Thesis", 100)]
    hdr_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=_BRAND_BLUE)
    for i, (label, width) in enumerate(cols, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    body = Font(name="Arial", size=9)
    sym_font = Font(name="Arial", size=9, bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for r_i, (_, row) in enumerate(df.iterrows(), start=DATA_START_ROW):
        vals = [row["Symbol"], row["Company"], row["CLD Thesis"]]
        for c_i, v in enumerate(vals, start=1):
            c = ws.cell(row=r_i, column=c_i, value=v)
            c.font = sym_font if c_i == 1 else body
            c.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=(c_i == 3),
            )
            c.border = Border(bottom=thin)
        # No explicit row height — Excel auto-fits the wrapped thesis text.

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"


def build_strategy_xlsx(
    *,
    strategy_code,
    tam_df,
    price_data,
    notion_data,
    div_data,
    warbook_data,
    fish_data,
    fish_history,
    as_of_date=None,
    logo_path=None,
):
    """
    Build a single workbook containing all four warbook tabs for the given
    strategy. Returns a BytesIO ready to hand to st.download_button.

    All data parameters are pre-fetched dicts from the same calls used by
    warbook_tab.py — no duplicated network/disk I/O.

    Args:
        strategy_code:    "QDVD" | "DAC" | "SMID" | "OR"
        tam_df:           DataFrame from get_holdings_for_strategy(...)
        price_data:       dict from fetch_batch_prices(tickers)
        notion_data:      dict from fetch_notion_metrics()
        div_data:         dict from get_batch_dividend_details(tickers)
        warbook_data:     dict from fetch_warbook_metrics_batch(tickers)
        fish_data:        dict { ticker: fish_metrics }
        fish_history:     dict { ticker: { year: dividend } }
        as_of_date:       date — printed in the title block; defaults to today
        logo_path:        path to MCP logo PNG; defaults to assets/M__Vector_.png

    Returns:
        BytesIO of the generated xlsx
    """
    if strategy_code not in WARBOOK_STRATEGIES:
        raise ValueError(
            f"{strategy_code} is not a warbook strategy. "
            f"Valid: {sorted(WARBOOK_STRATEGIES)}"
        )

    as_of_date = as_of_date or _date.today()
    logo_path = logo_path or _DEFAULT_LOGO
    strat_name = STRATEGY_DISPLAY_NAMES.get(strategy_code, strategy_code)

    wb = Workbook()
    # Remove default sheet — we'll add named ones
    wb.remove(wb.active)

    # ── Tab 1: Strategy Overview ───────────────────────────────────────
    df_ov = _build_overview_rows(tam_df, price_data, notion_data, div_data)
    ws = wb.create_sheet("Strategy Overview")
    if not df_ov.empty:
        _write_tab(ws, "overview", OVERVIEW_COLS, df_ov,
                   strategy_code, strat_name, as_of_date, logo_path)

    # ── Tab 1b: CLD Thesis Notes (companion to Strategy Overview) ───────
    df_notes = _build_cld_notes_rows(tam_df, notion_data)
    if not df_notes.empty:
        ws = wb.create_sheet("CLD Thesis Notes")
        _write_cld_notes_tab(ws, df_notes, strategy_code, strat_name,
                             as_of_date, logo_path)

    # ── Tab 2: QDG Characteristics ─────────────────────────────────────
    df_qdg = _build_qdg_rows(tam_df, price_data, notion_data, warbook_data,
                              fish_data, fish_history)
    ws = wb.create_sheet("QDG Characteristics")
    if not df_qdg.empty:
        _write_tab(ws, "qdg", QDG_COLS, df_qdg,
                   strategy_code, strat_name, as_of_date, logo_path)

    # ── Tab 3: Risk Correlation ────────────────────────────────────────
    df_risk = _build_risk_rows(tam_df, price_data, notion_data, warbook_data)
    ws = wb.create_sheet("Risk Correlation")
    if not df_risk.empty:
        _write_tab(ws, "risk", RISK_COLS, df_risk,
                   strategy_code, strat_name, as_of_date, logo_path)

    # ── Tab 4: Attribution ─────────────────────────────────────────────
    df_attr = _build_attribution_rows(tam_df, price_data, warbook_data)
    ws = wb.create_sheet("Attribution")
    if not df_attr.empty:
        _write_tab(ws, "attribution", ATTRIBUTION_COLS, df_attr,
                   strategy_code, strat_name, as_of_date, logo_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_single_tab_xlsx(
    *,
    tab_key,                # "overview" | "qdg" | "risk" | "attribution"
    strategy_code,
    tam_df,
    price_data,
    notion_data=None,
    div_data=None,
    warbook_data=None,
    fish_data=None,
    fish_history=None,
    as_of_date=None,
    logo_path=None,
):
    """
    Build a single-tab workbook. Useful for the per-tab download buttons.

    The data dicts that aren't needed for the requested tab can be omitted
    (defaults to empty dict).
    """
    if strategy_code not in WARBOOK_STRATEGIES:
        raise ValueError(
            f"{strategy_code} is not a warbook strategy. "
            f"Valid: {sorted(WARBOOK_STRATEGIES)}"
        )
    if tab_key not in ("overview", "qdg", "risk", "attribution"):
        raise ValueError(f"Unknown tab_key: {tab_key}")

    notion_data = notion_data or {}
    div_data = div_data or {}
    warbook_data = warbook_data or {}
    fish_data = fish_data or {}
    fish_history = fish_history or {}
    as_of_date = as_of_date or _date.today()
    logo_path = logo_path or _DEFAULT_LOGO
    strat_name = STRATEGY_DISPLAY_NAMES.get(strategy_code, strategy_code)

    if tab_key == "overview":
        df = _build_overview_rows(tam_df, price_data, notion_data, div_data)
        cols = OVERVIEW_COLS
        sheet_name = "Strategy Overview"
    elif tab_key == "qdg":
        df = _build_qdg_rows(tam_df, price_data, notion_data, warbook_data,
                              fish_data, fish_history)
        cols = QDG_COLS
        sheet_name = "QDG Characteristics"
    elif tab_key == "risk":
        df = _build_risk_rows(tam_df, price_data, notion_data, warbook_data)
        cols = RISK_COLS
        sheet_name = "Risk Correlation"
    else:  # attribution
        df = _build_attribution_rows(tam_df, price_data, warbook_data)
        cols = ATTRIBUTION_COLS
        sheet_name = "Attribution"

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    if not df.empty:
        _write_tab(ws, tab_key, cols, df,
                   strategy_code, strat_name, as_of_date, logo_path)

    # The Strategy Overview download carries its companion thesis-notes
    # sheet, mirroring the on-screen panel (and the legacy warbook's
    # hand-typed column U — now sourced from Notion, on the right rows).
    if tab_key == "overview" and notion_data:
        df_notes = _build_cld_notes_rows(tam_df, notion_data)
        if not df_notes.empty:
            ws_n = wb.create_sheet("CLD Thesis Notes")
            _write_cld_notes_tab(ws_n, df_notes, strategy_code, strat_name,
                                 as_of_date, logo_path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_filename(strategy_code, tab_key=None, as_of_date=None):
    """
    Standard filename convention.

    All tabs:    MCP_Warbook_{STRATEGY}_{YYYY-MM-DD}.xlsx
    Single tab:  MCP_Warbook_{STRATEGY}_{TabName}_{YYYY-MM-DD}.xlsx
    """
    as_of_date = as_of_date or _date.today()
    date_str = as_of_date.strftime("%Y-%m-%d")
    if tab_key is None:
        return f"MCP_Warbook_{strategy_code}_{date_str}.xlsx"
    tab_label = {
        "overview":     "StrategyOverview",
        "qdg":          "QDG",
        "risk":         "RiskCorrelation",
        "attribution":  "Attribution",
    }.get(tab_key, tab_key)
    return f"MCP_Warbook_{strategy_code}_{tab_label}_{date_str}.xlsx"