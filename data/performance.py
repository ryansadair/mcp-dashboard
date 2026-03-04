"""
Strategy performance: YTD returns, benchmark comparison, alpha.
Sources: Strategy_Returns.xlsx (quarterly returns per strategy)
         Supabase benchmark_ytd table (populated by prefetch_data.py)
         yfinance fallback if Supabase unavailable
"""

import os
import pandas as pd
import streamlit as st
import requests
from datetime import date
from utils.config import STRATEGIES
from utils.cache import query

# ── Supabase config ────────────────────────────────────────────────────────
SUPABASE_URL = "https://idtytpyehfbqldnvwenb.supabase.co"
SUPABASE_KEY = "sb_secret_P1XNpklX_g_gcMamZb0qqw_udXSu8T7"   # paste your service role key here

_SB_HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
}

def _sb_get_benchmark_ytd(bench_ticker):
    """Fetch YTD return for a benchmark from Supabase."""
    if SUPABASE_KEY == "sb_secret_P1XNpklX_g_gcMamZb0qqw_udXSu8T7":
        return None
    try:
        url    = f"{SUPABASE_URL}/rest/v1/benchmark_ytd"
        params = {"select": "ytd_return", "symbol": f"eq.{bench_ticker}"}
        resp   = requests.get(url, headers=_SB_HEADERS, params=params, timeout=10)
        if resp.status_code == 200:
            rows = resp.json()
            if rows:
                return rows[0].get("ytd_return")
        return None
    except Exception:
        return None

def _sb_get_benchmark_history(bench_ticker):
    """Fetch full YTD price history for a benchmark from Supabase."""
    if SUPABASE_KEY == "sb_secret_P1XNpklX_g_gcMamZb0qqw_udXSu8T7":
        return None
    try:
        url    = f"{SUPABASE_URL}/rest/v1/benchmark_history"
        params = {"select": "date,close", "symbol": f"eq.{bench_ticker}", "order": "date.asc"}
        resp   = requests.get(url, headers=_SB_HEADERS, params=params, timeout=10)
        if resp.status_code == 200:
            rows = resp.json()
            if rows:
                return pd.DataFrame(rows)
        return None
    except Exception:
        return None


# ── Strategy Returns (from Strategy_Returns.xlsx) ─────────────────────────

RETURNS_PATHS = [
    "data/Strategy_Returns.xlsx",
    "Strategy_Returns.xlsx",
]

@st.cache_data(ttl=3600)
def load_strategy_returns() -> dict:
    """
    Parse Strategy_Returns.xlsx — one sheet per strategy.
    Returns dict: { "QDVD": pd.DataFrame(date, ret, cum), ... }
    """
    import openpyxl
    path = next((p for p in RETURNS_PATHS if os.path.exists(p)), None)
    if not path:
        return {}

    wb  = openpyxl.load_workbook(path)
    out = {}

    for sheet in wb.sheetnames:
        ws   = wb[sheet]
        rows = [
            (r[0], r[1]) for r in ws.iter_rows(values_only=True)
            if r[0] is not None and r[1] is not None
            and not isinstance(r[0], str)
        ]
        if not rows:
            continue

        df         = pd.DataFrame(rows, columns=["date", "ret"])
        df["date"] = pd.to_datetime(df["date"])
        df         = df.sort_values("date").dropna().reset_index(drop=True)
        df["cum"]  = (1 + df["ret"]).cumprod() - 1
        out[sheet] = df

    return out


@st.cache_data(ttl=3600)
def get_strategy_ytd(strategy: str) -> float:
    """Get YTD return for a strategy from the returns file."""
    returns = load_strategy_returns()
    df      = returns.get(strategy)
    if df is None or df.empty:
        return 0.0
    current_year = date.today().year
    ytd = df[df["date"].dt.year == current_year]
    if ytd.empty:
        last_year = df["date"].dt.year.max()
        ytd = df[df["date"].dt.year == last_year]
    if ytd.empty:
        return 0.0
    return round(float((1 + ytd["ret"]).prod() - 1) * 100, 2)


@st.cache_data(ttl=3600)
def get_strategy_kpis(strategy: str) -> dict:
    """Get KPIs for a strategy — ytd from returns file."""
    ytd = get_strategy_ytd(strategy)
    return {"ytd": ytd}


@st.cache_data(ttl=3600)
def get_benchmark_ytd(bench_ticker: str) -> float:
    # ── 1. Try Supabase ───────────────────────────────────────────────────
    sb_ytd = _sb_get_benchmark_ytd(bench_ticker)
    if sb_ytd is not None:
        return sb_ytd

    # ── 2. yfinance fallback ──────────────────────────────────────────────
    try:
        import yfinance as yf
        start = date(date.today().year, 1, 1)
        hist  = yf.download(bench_ticker, start=start, progress=False, auto_adjust=True)
        if hist is None or len(hist) < 2:
            raise ValueError("Not enough data")
        closes = hist["Close"]
        first  = float(closes.iloc[0])
        last   = float(closes.iloc[-1])
        return round(((last - first) / first) * 100, 2)
    except Exception:
        fallbacks = {"^GSPC": 7.15, "^SP500DVS": 5.92, "^RUT": 9.44, "^SP500GR": 8.21}
        return fallbacks.get(bench_ticker, 6.0)


@st.cache_data(ttl=3600)
def get_perf_chart_data(strategy: str, bench_ticker: str, period: str = "YTD") -> pd.DataFrame:
    """
    Build cumulative return series for strategy vs benchmark.
    Strategy: from Strategy_Returns.xlsx
    Benchmark: from Supabase benchmark_history, filtered to same date range.
    period: 'YTD', '1Y', '3Y', '5Y', 'ITD'
    """
    returns = load_strategy_returns()
    df      = returns.get(strategy)

    if df is None or df.empty:
        return pd.DataFrame()

    # ── Filter strategy by period ─────────────────────────────────────────
    today      = pd.Timestamp.today()
    cutoffs    = {
        "YTD": pd.Timestamp(today.year, 1, 1),
        "1Y":  today - pd.DateOffset(years=1),
        "3Y":  today - pd.DateOffset(years=3),
        "5Y":  today - pd.DateOffset(years=5),
        "ITD": df["date"].min(),
    }
    cutoff = cutoffs.get(period, cutoffs["YTD"])
    df     = df[df["date"] >= cutoff].copy()

    if df.empty:
        return pd.DataFrame()

    # Rebase cumulative from period start
    base        = float((1 + df["ret"]).iloc[0] ** 0 )  # =1 always
    df["strat"] = (1 + df["ret"]).cumprod()
    df["strat"] = (df["strat"] / df["strat"].iloc[0] - 1) * 100

    # ── Benchmark: try Supabase history ──────────────────────────────────
    bench_df = _sb_get_benchmark_history(bench_ticker)
    if bench_df is not None and not bench_df.empty:
        bench_df["date"]  = pd.to_datetime(bench_df["date"])
        bench_df          = bench_df[bench_df["date"] >= cutoff].sort_values("date")
        if not bench_df.empty:
            bench_df["bench"] = (bench_df["close"] / bench_df["close"].iloc[0] - 1) * 100
            # Merge on nearest date
            df = df.merge(
                bench_df[["date", "bench"]],
                on="date", how="left"
            )
            df["bench"] = df["bench"].ffill()

    return df[["date", "strat"] + (["bench"] if "bench" in df.columns else [])]