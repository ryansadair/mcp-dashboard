"""
Martin Capital Partners — Tamarac File Auto-Detector
data/tamarac_detector.py

Scans configured directories for the newest Tamarac holdings export.
Returns file metadata (path, modified time, as-of date) so the dashboard
can detect when a newer file has been committed or dropped in.

Usage in 1_Dashboard.py:
    from data.tamarac_detector import find_best_tamarac_file, get_tamarac_status

    status = get_tamarac_status()
    if status["found"]:
        tamarac_parsed = _load_tamarac(status["path"])
    if status["stale"]:
        st.warning(f"Tamarac data is {status['age_days']} days old — consider uploading a fresh export.")
"""

import os
import glob
from pathlib import Path
from datetime import datetime, timedelta

import streamlit as st

# ── Directory resolution ───────────────────────────────────────────────────
_THIS_DIR = Path(__file__).parent
_PROJECT_ROOT = _THIS_DIR.parent

# Where to look for Tamarac files, in priority order.
# First match wins.  On Streamlit Cloud only data/ will exist;
# locally Ryan's import folder is checked first.
_SCAN_DIRS = [
    _THIS_DIR / "tamarac_imports",            # data/tamarac_imports/
    _THIS_DIR,                                 # data/
    _PROJECT_ROOT,                             # project root
]

# Also honour the explicit path from config, if available
try:
    from utils.config import TAMARAC_WATCH_FOLDER
    if TAMARAC_WATCH_FOLDER:
        _SCAN_DIRS.insert(0, Path(TAMARAC_WATCH_FOLDER))
except ImportError:
    pass

# Filename patterns to match (case-insensitive via glob)
_FILE_PATTERNS = [
    "Tamarac_Holdings*.xlsx",
    "tamarac_holdings*.xlsx",
    "Tamarac_Export*.xlsx",
    "tamarac_export*.xlsx",
    "Holdings_Export*.xlsx",
]

# How old (in days) before we flag the data as stale
STALE_THRESHOLD_DAYS = 7


# ── As-of-date reader ─────────────────────────────────────────────────────

def _read_as_of_date(path):
    """
    Read the 'As of Date' from inside a Tamarac Excel file (cell A2 of
    the first sheet).  Returns a datetime, or None on failure.

    This is more reliable than os.path.getmtime() because OneDrive syncs
    can update the filesystem modification time without changing the file
    content.  The 'As of Date' only changes when Ryan exports a new file
    from Tamarac.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        val = ws.cell(2, 1).value
        wb.close()
        if isinstance(val, datetime):
            return val
        # Try parsing string dates
        if val:
            return datetime.strptime(str(val)[:10], "%Y-%m-%d")
    except Exception:
        pass
    return None


# ── Core lookup ────────────────────────────────────────────────────────────

def _scan_for_files():
    """
    Scan all configured directories for Tamarac export files.
    Returns list of (path_str, mtime_float) sorted newest-first.
    """
    candidates = []
    seen_paths = set()

    for scan_dir in _SCAN_DIRS:
        scan_dir = Path(scan_dir)
        if not scan_dir.is_dir():
            continue
        for pattern in _FILE_PATTERNS:
            for match in scan_dir.glob(pattern):
                real = str(match.resolve())
                if real not in seen_paths:
                    seen_paths.add(real)
                    candidates.append((str(match), os.path.getmtime(str(match))))

    # Sort newest first
    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates


def find_best_tamarac_file():
    """
    Return the path to the newest Tamarac holdings file, or None.
    """
    files = _scan_for_files()
    if files:
        return files[0][0]
    return None


def find_all_tamarac_files():
    """
    Return list of all found Tamarac files as dicts:
    [{"path": str, "modified": datetime, "as_of_date": datetime|None,
      "size_kb": float, "filename": str}, ...]
    Sorted newest-first by as_of_date (falls back to mtime if unreadable).
    """
    results = []
    for path_str, mtime in _scan_for_files():
        as_of = _read_as_of_date(path_str)
        results.append({
            "path": path_str,
            "filename": os.path.basename(path_str),
            "modified": datetime.fromtimestamp(mtime),
            "as_of_date": as_of,
            "size_kb": round(os.path.getsize(path_str) / 1024, 1),
        })
    # Re-sort by as_of_date (prefer internal date), falling back to mtime
    results.sort(
        key=lambda x: x["as_of_date"] or x["modified"],
        reverse=True,
    )
    return results


@st.cache_data(ttl=60, show_spinner=False)
def get_tamarac_status():
    """
    Dashboard-ready status dict.  Cached for 60s so we don't hit the
    filesystem on every Streamlit rerun.

    Returns:
        {
            "found":      bool,
            "path":       str | None,
            "filename":   str | None,
            "modified":   datetime | None,
            "age_days":   int,
            "stale":      bool,          # True if > STALE_THRESHOLD_DAYS old
            "all_files":  list[dict],    # every file found, newest first
        }
    """
    files = find_all_tamarac_files()

    if not files:
        return {
            "found": False,
            "path": None,
            "filename": None,
            "modified": None,
            "age_days": 0,
            "stale": True,
            "all_files": [],
        }

    best = files[0]
    # Use the internal "As of Date" from the Excel file for age calculation.
    # This is immune to OneDrive syncs bumping the filesystem mtime.
    # Fall back to mtime if the internal date can't be read.
    effective_date = best.get("as_of_date") or best["modified"]
    age = (datetime.now() - effective_date).days

    return {
        "found": True,
        "path": best["path"],
        "filename": best["filename"],
        "modified": best["modified"],
        "as_of_date": best.get("as_of_date"),
        "age_days": age,
        "stale": age > STALE_THRESHOLD_DAYS,
        "all_files": files,
    }


# ── Dashboard banner helper ────────────────────────────────────────────────

def render_tamarac_status_banner():
    """
    Render a small status banner in the dashboard showing Tamarac file info.
    Call this right after loading Tamarac data in 1_Dashboard.py.
    """
    status = get_tamarac_status()

    if not status["found"]:
        st.warning(
            "⚠️ **No Tamarac file found.** "
            "Place `Tamarac_Holdings.xlsx` in the `data/` folder and commit to git.",
            icon="📂"
        )
        return status

    # Show the internal "As of Date" if available, otherwise fall back to mtime
    display_date = status.get("as_of_date") or status["modified"]
    mod_str = display_date.strftime("%b %d, %Y")
    age = status["age_days"]

    if status["stale"]:
        st.warning(
            f"📂 **Tamarac data is {age} days old** — "
            f"`{status['filename']}` as-of {mod_str}. "
            f"Consider uploading a fresh export.",
            icon="⏰"
        )
    else:
        # Subtle info line — not a warning
        st.markdown(
            f"<div style='font-size:11px;color:rgba(255,255,255,0.3);padding:2px 0 6px 0;'>"
            f"📂 Tamarac: <span style=\"color:rgba(86,149,66,0.7);\">{status['filename']}</span>"
            f" · as-of {mod_str}"
            f"{'  · <span style=\"color:rgba(201,168,76,0.7);\">(' + str(age) + 'd ago)</span>' if age > 0 else ''}"
            f"</div>",
            unsafe_allow_html=True,
        )

    return status