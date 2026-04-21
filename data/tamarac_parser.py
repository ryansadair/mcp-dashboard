"""
Tamarac Holdings Excel Parser
Reads multi-sheet Tamarac export files and returns structured holdings data.
Handles the =\"value\" quoting format from Tamarac exports.
"""
import pandas as pd
import openpyxl
import os
import sqlite3
from datetime import datetime

# Strategy display names
STRATEGY_NAMES = {
    "QDVD": "Quality Dividend",
    "SMID": "Quality SMID Dividend",
    "DAC": "Quality All-Cap Dividend",
    "OR": "Oregon Dividend",
    "DCP": "Dividend Core Plus",
}

STRATEGY_COLORS = {
    "QDVD": "#569542",
    "SMID": "#C9A84C",
    "DAC": "#07415A",
    "OR": "#569542",
    "DCP": "#07415A",
}

# Benchmarks for each strategy
STRATEGY_BENCHMARKS = {
    "QDVD": {"name": "S&P 500", "ticker": "^GSPC"},
    "SMID": {"name": "Russell 2500", "ticker": "^SP600"},  # proxy
    "DAC":  {"name": "S&P 500", "ticker": "^GSPC"},
    "OR":   {"name": "S&P 500", "ticker": "^GSPC"},
    "DCP":  {"name": "S&P 500", "ticker": "^GSPC"},
}


def clean_tamarac_value(val):
    """Strip the =\"...\" quoting that Tamarac exports use."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.startswith('="') and s.endswith('"'):
        return s[2:-1]
    if s.startswith('*'):
        return s  # strategy name field
    return s


def _parse_sheet_rows(rows):
    """
    Parse one Tamarac sheet (already materialized rows) into a DataFrame with
    standardized column names. Used by both the auto-pulled file and the
    manual-fallback file, which share the same multi-sheet format.

    Returns an empty DataFrame if the sheet has no data rows.
    """
    if len(rows) < 2:
        return pd.DataFrame()

    headers_raw = rows[0]
    headers = [clean_tamarac_value(h).lower().replace(" ", "_") for h in headers_raw]

    data = []
    for row in rows[1:]:
        record = {}
        for j, val in enumerate(row):
            col = headers[j] if j < len(headers) else f"col_{j}"
            if col == "as_of_date":
                record[col] = val if isinstance(val, datetime) else None
            elif col == "weight":
                try:
                    str_val = str(val).strip() if val else ""
                    if str_val.endswith("%"):
                        record[col] = float(str_val.rstrip("%")) / 100
                    else:
                        record[col] = float(val) if val else 0.0
                except (ValueError, TypeError):
                    record[col] = 0.0
            elif col in ("quantity", "yield_at_cost", "current_yield",
                         "unit_cost", "cost_basis", "value", "price",
                         "annual_income", "cumulative_income"):
                try:
                    str_val = str(val).strip() if val else ""
                    if str_val.endswith("%"):
                        record[col] = float(str_val.rstrip("%")) / 100
                    else:
                        record[col] = float(val) if val else 0.0
                except (ValueError, TypeError):
                    record[col] = 0.0
            else:
                record[col] = clean_tamarac_value(val)
        data.append(record)

    df = pd.DataFrame(data)

    # Rename columns to standard names
    col_map = {
        "as_of_date": "as_of_date",
        "data_for": "strategy_raw",
        "weight": "weight",
        "symbol": "symbol",
        "cusip": "cusip",
        "description": "description",
        "quantity": "quantity",
        "open_date": "open_date",
        "value": "value",
        "price": "price",
        "unit_cost": "unit_cost",
        "cost_basis": "cost_basis",
        "annual_income": "annual_income",
        "cumulative_income": "cumulative_income",
        "yield_at_cost": "yield_at_cost",
        "current_yield": "current_yield",
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
    return df


# Per-ticker cost/yield fields that the manual file provides and the
# auto-pulled API export does not. Merged by (strategy_code, symbol).
_MANUAL_MERGE_FIELDS = [
    "unit_cost",
    "cost_basis",
    "annual_income",
    "cumulative_income",
    "yield_at_cost",
    "current_yield",
]


def _build_manual_lookup(manual_filepath):
    """
    Parse Tamarac_Holdings_Manual.xlsx (if present) and return a dict keyed by
    (strategy_code, symbol) -> {unit_cost, cost_basis, ...}

    Returns {} if file is missing or unreadable; callers should treat an empty
    lookup as "no manual data available" and proceed without merging.
    """
    if not manual_filepath or not os.path.exists(manual_filepath):
        return {}
    try:
        wb = openpyxl.load_workbook(manual_filepath, read_only=True)
    except Exception:
        return {}

    lookup = {}
    for sheet_name in wb.sheetnames:
        try:
            rows = list(wb[sheet_name].iter_rows(values_only=True))
        except Exception:
            continue
        df = _parse_sheet_rows(rows)
        if df.empty or "symbol" not in df.columns:
            continue
        for _, r in df.iterrows():
            sym = str(r.get("symbol") or "").strip().upper()
            if not sym:
                continue
            key = (sheet_name, sym)
            lookup[key] = {
                f: r.get(f, 0) for f in _MANUAL_MERGE_FIELDS if f in df.columns
            }
    wb.close()
    return lookup


def _apply_manual_lookup(df, strategy_code, lookup):
    """
    For each row in df, if (strategy_code, symbol) is in the lookup, copy
    the manual cost/yield fields into the row. Creates the columns if they
    don't already exist (the auto-pulled file won't have them).
    """
    if df.empty or not lookup:
        return df

    # Ensure all target columns exist so downstream .get() calls see real
    # floats (0.0) rather than KeyError / NaN surprises.
    for field in _MANUAL_MERGE_FIELDS:
        if field not in df.columns:
            df[field] = 0.0

    def _fill(row):
        sym = str(row.get("symbol") or "").strip().upper()
        manual = lookup.get((strategy_code, sym))
        if not manual:
            return row
        for field, val in manual.items():
            # Only overwrite when the manual value is meaningful (non-zero).
            # Lets the auto-pull win in the rare case it actually provides
            # one of these fields and the manual file has it as 0.
            if val is not None and val != 0:
                row[field] = val
        return row

    return df.apply(_fill, axis=1)


def parse_tamarac_excel(filepath):
    """
    Parse a Tamarac Holdings Excel export.

    Returns dict: {strategy_code: DataFrame} where each DF has columns:
        as_of_date, strategy_raw, weight, symbol, cusip, description, quantity,
        unit_cost, cost_basis, annual_income, cumulative_income,
        yield_at_cost, current_yield, is_cash

    Data source strategy:
      1. Auto-pulled file (filepath) — fresh daily positions/weights/quantities
         but lacks cost basis and yield data (Tamarac API template 41 limit).
      2. Manual file (Tamarac_Holdings_Manual.xlsx in same folder, if present) —
         updated on a slower cadence. We use it purely as a per-(strategy,ticker)
         lookup to fill in unit_cost, cost_basis, annual_income, cumulative_income,
         yield_at_cost, current_yield. Positions/weights always come from (1).

      If the manual file is missing, those six fields will be 0.0 and the
      dashboard will render them as em-dashes where appropriate.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        df = _parse_sheet_rows(rows)
        if df.empty:
            continue

        # Filter out CASH rows for holdings display (keep for weight calc)
        if "symbol" in df.columns:
            df["is_cash"] = df["symbol"].astype(str).str.upper() == "CASH"
        else:
            df["is_cash"] = False

        result[sheet_name] = df

    wb.close()

    # ── Merge in manual cost/yield data if the fallback file exists ───────
    # Looks for Tamarac_Holdings_Manual.xlsx in the SAME directory as the
    # auto-pulled file. The manual file is optional — absent = zero fields.
    folder = os.path.dirname(os.path.abspath(filepath))
    manual_path = os.path.join(folder, "Tamarac_Holdings_Manual.xlsx")
    lookup = _build_manual_lookup(manual_path)
    if lookup:
        for strategy_code, df in result.items():
            result[strategy_code] = _apply_manual_lookup(df, strategy_code, lookup)

    return result


def get_holdings_for_strategy(parsed_data, strategy_code, include_cash=False):
    """Get holdings DataFrame for a specific strategy, optionally excluding cash."""
    if strategy_code not in parsed_data:
        return pd.DataFrame()

    df = parsed_data[strategy_code].copy()
    if not include_cash:
        df = df[~df["is_cash"]].copy()

    # Convert weight from decimal to percentage for display
    df["weight_pct"] = df["weight"] * 100

    return df.sort_values("weight", ascending=False).reset_index(drop=True)


def get_cash_weight(parsed_data, strategy_code):
    """Get cash weight for a strategy."""
    if strategy_code not in parsed_data:
        return 0.0
    df = parsed_data[strategy_code]
    cash_rows = df[df["is_cash"]]
    if len(cash_rows) > 0:
        return float(cash_rows["weight"].iloc[0]) * 100
    return 0.0


def get_all_unique_tickers(parsed_data):
    """Get sorted list of all unique non-cash tickers across all strategies."""
    tickers = set()
    for strategy_code, df in parsed_data.items():
        non_cash = df[~df["is_cash"]]
        tickers.update(non_cash["symbol"].str.upper().tolist())
    return sorted(tickers)


def save_holdings_to_db(parsed_data, db_path="data/portfolio.db"):
    """Cache parsed holdings to SQLite."""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)

    for strategy_code, df in parsed_data.items():
        df_save = df.copy()
        df_save["strategy_code"] = strategy_code
        if "as_of_date" in df_save.columns:
            df_save["as_of_date"] = df_save["as_of_date"].astype(str)
        df_save.to_sql("holdings", conn, if_exists="append", index=False)

    conn.close()


def get_as_of_date(parsed_data):
    """Get the as-of date from the first available strategy."""
    for code, df in parsed_data.items():
        if "as_of_date" in df.columns and len(df) > 0:
            val = df["as_of_date"].iloc[0]
            if isinstance(val, datetime):
                return val
    return None