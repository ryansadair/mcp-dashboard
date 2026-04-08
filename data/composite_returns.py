"""
Martin Capital Partners — Composite Returns Parser
Reads Composite Returns.xls from OneDrive and extracts monthly/quarterly
returns, cumulative values, and benchmark data for each equity strategy.

Source file: Composite Returns.xls (Tamarac/Black Diamond export)
OneDrive path: C:/Users/RyanAdair/Martin Capital Partners LLC/
               Eugene - Documents/Performance/Composite Returns/Composite Returns.xls

Strategies parsed: QDVD, SMID, DAC, OR (equity composites only)
"""

import os
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)

# ── File Location ───────────────────────────────────────────────────────────
# OneDrive paths (Windows / local dev) — check both extensions
ONEDRIVE_DIR = (
    r"C:\Users\RyanAdair\Martin Capital Partners LLC"
    r"\Eugene - Documents\Performance\Composite Returns"
)

# Fallback: check data/ folder (for Streamlit Cloud if file is committed)
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_THIS_DIR)


# ── Strategy Block Definitions ──────────────────────────────────────────────
# Each strategy composite is a block in the "Composite Returns" sheet.
# start_row = row of the strategy label (e.g., "QDVD Composite")
# header_row = start_row + 1 (contains column headers)
# data starts at header_row + 1
#
# Column layout (consistent across all blocks):
#   0: Period Ending (Excel date serial)
#   1: Ending Value (AUM)
#   2: # of Accounts
#   3: Net Quarterly Return
#   4: Net Monthly Return
#   5: Cumulative Net Return
#   6: Cumulative Net Value
#   7: Gross Quarterly Return
#   8: Gross Monthly Return
#   9: Cumulative Gross Return
#  10: Cumulative Gross Value
#  11+: Benchmark columns (vary by strategy)

COMPOSITE_BLOCKS = {
    "QDVD": {
        "label_row": 2,
        "benchmarks": {
            "primary": {"name": "S&P 500 High Dividend", "qtr_col": 15, "mo_col": 16, "cum_col": 17, "val_col": 18},
            "secondary": {"name": "S&P 500", "qtr_col": 11, "mo_col": 12, "cum_col": 13, "val_col": 14},
        },
    },
    "SMID": {
        "label_row": 159,
        "benchmarks": {
            "primary": {"name": "S&P Mid Cap 400", "qtr_col": 11, "mo_col": 12, "cum_col": 13, "val_col": 14},
            "secondary": {"name": "S&P 400 Aristocrats", "qtr_col": 15, "mo_col": 16, "cum_col": 17, "val_col": 18},
        },
    },
    "DAC": {
        "label_row": 296,
        "benchmarks": {
            "primary": {"name": "Russell 3000", "qtr_col": 11, "mo_col": 12, "cum_col": 13, "val_col": 14},
            "secondary": {"name": "Dow Jones Select Dividend", "qtr_col": 15, "mo_col": 16, "cum_col": 17, "val_col": 18},
        },
    },
    "OR": {
        "label_row": 421,
        "benchmarks": {
            "primary": {"name": "S&P 500", "qtr_col": 11, "mo_col": 12, "cum_col": 13, "val_col": 14},
            # OR only has one benchmark
        },
    },
}

# ── Period Returns Sheet Mapping ────────────────────────────────────────────
# Row indices in the "Period Returns" sheet for each strategy/benchmark
PERIOD_RETURNS_ROWS = {
    "QDVD": 5,
    "SMID": 6,
    "DAC": 7,
    "OR": 8,
    # Benchmarks
    "S&P 500": 12,
    "S&P 500 High Dividend": 11,
    "S&P Mid Cap 400": 14,
    "S&P 400 Aristocrats": 13,
    "Russell 3000": 16,
    "Dow Jones Select Dividend": 15,
}

# Column indices in Period Returns: {col: period_label}
PERIOD_COLS = {
    1: "QTD",
    2: "YTD",
    3: "1Y",
    4: "3Y",
    5: "5Y",
    6: "10Y",
    7: "Since Inception",
    8: "Since Inception (Ann.)",
}


# ── Helper Functions ────────────────────────────────────────────────────────

def _find_composite_file():
    """Locate Composite Returns file — checks .xlsx first, then .xls, in OneDrive then data/."""
    # Search order: newest format first, OneDrive first
    candidates = [
        os.path.join(ONEDRIVE_DIR, "Composite Returns.xlsx"),
        os.path.join(ONEDRIVE_DIR, "Composite Returns.xls"),
        os.path.join(_THIS_DIR, "Composite Returns.xlsx"),
        os.path.join(_THIS_DIR, "Composite Returns.xls"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            logger.info(f"Composite returns found: {path}")
            return path
    logger.warning("Composite Returns file not found in any expected location (.xlsx or .xls)")
    return None


def _excel_date_to_datetime(serial, datemode=0):
    """Convert Excel date serial number to Python datetime."""
    if isinstance(serial, datetime):
        return serial  # openpyxl already returns datetime objects
    try:
        return xlrd.xldate_as_datetime(serial, datemode)
    except Exception:
        return None


def _safe_float(val):
    """Convert cell value to float, returning NaN for blanks/errors."""
    if val == '' or val is None:
        return np.nan
    try:
        return float(val)
    except (ValueError, TypeError):
        return np.nan


class _OpenpyxlSheetWrapper:
    """
    Wraps an openpyxl worksheet to provide the same API as xlrd:
      ws.cell_value(row, col)  — 0-indexed
      ws.nrows
      ws.ncols
    This lets existing parsers work unchanged with both backends.
    """

    def __init__(self, ws):
        self._ws = ws
        self.nrows = ws.max_row or 0
        self.ncols = ws.max_column or 0

    def cell_value(self, row, col):
        """Return cell value using 0-indexed row/col (like xlrd)."""
        val = self._ws.cell(row=row + 1, column=col + 1).value
        return val if val is not None else ''


# ── Core Parser ─────────────────────────────────────────────────────────────

def _parse_composite_block(ws, strategy, datemode):
    """
    Parse a single composite block from the Composite Returns sheet.
    Returns a DataFrame with columns:
        date, aum, accounts,
        net_qtr, net_mo, cum_net, cum_net_val,
        gross_qtr, gross_mo, cum_gross, cum_gross_val,
        bench1_name, bench1_qtr, bench1_mo, bench1_cum, bench1_val,
        bench2_name, bench2_qtr, bench2_mo, bench2_cum, bench2_val
    """
    block = COMPOSITE_BLOCKS[strategy]
    header_row = block["label_row"] + 1
    data_start = header_row + 1
    benchmarks = block["benchmarks"]

    rows = []
    for r in range(data_start, ws.nrows):
        date_val = ws.cell_value(r, 0)
        if date_val == '' or date_val is None:
            break  # End of block

        dt = _excel_date_to_datetime(date_val, datemode)
        if dt is None:
            break

        row = {
            "date": dt,
            "aum": _safe_float(ws.cell_value(r, 1)),
            "accounts": _safe_float(ws.cell_value(r, 2)),
            "net_qtr": _safe_float(ws.cell_value(r, 3)),
            "net_mo": _safe_float(ws.cell_value(r, 4)),
            "cum_net": _safe_float(ws.cell_value(r, 5)),
            "cum_net_val": _safe_float(ws.cell_value(r, 6)),
            "gross_qtr": _safe_float(ws.cell_value(r, 7)),
            "gross_mo": _safe_float(ws.cell_value(r, 8)),
            "cum_gross": _safe_float(ws.cell_value(r, 9)),
            "cum_gross_val": _safe_float(ws.cell_value(r, 10)),
        }

        # Primary benchmark
        p = benchmarks.get("primary", {})
        row["bench1_name"] = p.get("name", "")
        row["bench1_qtr"] = _safe_float(ws.cell_value(r, p["qtr_col"])) if "qtr_col" in p else np.nan
        row["bench1_mo"] = _safe_float(ws.cell_value(r, p["mo_col"])) if "mo_col" in p else np.nan
        row["bench1_cum"] = _safe_float(ws.cell_value(r, p["cum_col"])) if "cum_col" in p else np.nan
        row["bench1_val"] = _safe_float(ws.cell_value(r, p["val_col"])) if "val_col" in p else np.nan

        # Secondary benchmark (if exists)
        s = benchmarks.get("secondary", {})
        row["bench2_name"] = s.get("name", "")
        row["bench2_qtr"] = _safe_float(ws.cell_value(r, s["qtr_col"])) if "qtr_col" in s else np.nan
        row["bench2_mo"] = _safe_float(ws.cell_value(r, s["mo_col"])) if "mo_col" in s else np.nan
        row["bench2_cum"] = _safe_float(ws.cell_value(r, s["cum_col"])) if "cum_col" in s else np.nan
        row["bench2_val"] = _safe_float(ws.cell_value(r, s["val_col"])) if "val_col" in s else np.nan

        rows.append(row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    return df


def _parse_period_returns(ws, datemode):
    """
    Parse the Period Returns summary sheet.
    Returns a dict: {strategy_or_benchmark_name: {period_label: value}}
    """
    result = {}
    for name, row_idx in PERIOD_RETURNS_ROWS.items():
        if row_idx >= ws.nrows:
            continue
        periods = {}
        for col, label in PERIOD_COLS.items():
            val = ws.cell_value(row_idx, col)
            if val != '' and val != '-' and val is not None:
                try:
                    periods[label] = float(val)
                except (ValueError, TypeError):
                    pass
        if periods:
            result[name] = periods
    return result


def _parse_annual_returns(ws, datemode):
    """
    Parse the annual returns table from Period Returns sheet.
    Returns a DataFrame with columns: Year, QDVD, SMID, DAC, OR, + benchmarks
    """
    # Annual returns start at row 29 (header), data from row 30
    header_row = 29
    if header_row >= ws.nrows:
        return pd.DataFrame()

    # Read column headers
    col_names = {}
    for c in range(ws.ncols):
        v = ws.cell_value(header_row, c)
        if v != '':
            col_names[c] = str(v)

    rows = []
    for r in range(header_row + 1, ws.nrows):
        year_val = ws.cell_value(r, 0)
        if year_val == '' or year_val is None:
            continue
        try:
            year = int(float(year_val))
        except (ValueError, TypeError):
            continue

        row = {"Year": year}
        for c, name in col_names.items():
            if c == 0:
                continue
            val = ws.cell_value(r, c)
            if val != '' and val != '-' and val is not None:
                try:
                    row[name] = float(val)
                except (ValueError, TypeError):
                    pass
        rows.append(row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows).sort_values("Year", ascending=False).reset_index(drop=True)
    return df


# ── Public API ──────────────────────────────────────────────────────────────

def load_composite_data():
    """
    Load all composite return data from Composite Returns.xls.

    Returns:
        dict with keys:
            "available": bool — whether file was found and parsed
            "as_of": datetime — reporting date
            "file_path": str — path used
            "composites": {strategy: DataFrame} — monthly/quarterly data per strategy
            "period_returns": {name: {period: value}} — summary period returns
            "annual_returns": DataFrame — calendar year returns
            "error": str or None — error message if failed
    """
    result = {
        "available": False,
        "as_of": None,
        "file_path": None,
        "composites": {},
        "period_returns": {},
        "annual_returns": pd.DataFrame(),
        "error": None,
    }

    if xlrd is None and openpyxl is None:
        result["error"] = "Neither xlrd nor openpyxl installed — cannot read Excel files"
        logger.error(result["error"])
        return result

    filepath = _find_composite_file()
    if filepath is None:
        result["error"] = (
            "Composite Returns file not found. "
            "Expected at OneDrive path or in data/ folder (.xls or .xlsx)."
        )
        return result

    is_xlsx = filepath.lower().endswith(".xlsx")

    if is_xlsx and openpyxl is None:
        result["error"] = "openpyxl not installed — cannot read .xlsx files"
        logger.error(result["error"])
        return result
    if not is_xlsx and xlrd is None:
        result["error"] = "xlrd not installed — cannot read .xls files"
        logger.error(result["error"])
        return result

    try:
        if is_xlsx:
            # ── openpyxl path (.xlsx) ──
            wb_ox = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            result["file_path"] = filepath
            datemode = 0  # not used for openpyxl (returns native datetimes)

            ws_main = _OpenpyxlSheetWrapper(wb_ox["Composite Returns"])

            # Parse "as of" date from row 1, col 1
            as_of_val = ws_main.cell_value(1, 1)
            if isinstance(as_of_val, datetime):
                result["as_of"] = as_of_val
            elif as_of_val and isinstance(as_of_val, (int, float)):
                try:
                    result["as_of"] = xlrd.xldate_as_datetime(as_of_val, 0) if xlrd else None
                except Exception:
                    pass

            # Parse each equity composite block
            for strategy in ["QDVD", "SMID", "DAC", "OR"]:
                try:
                    df = _parse_composite_block(ws_main, strategy, datemode)
                    if not df.empty:
                        result["composites"][strategy] = df
                        logger.info(f"  {strategy}: {len(df)} rows, "
                                    f"{df['date'].min().strftime('%Y-%m')} to "
                                    f"{df['date'].max().strftime('%Y-%m')}")
                except Exception as e:
                    logger.error(f"  Error parsing {strategy}: {e}")

            # Parse Period Returns sheet
            try:
                ws_period = _OpenpyxlSheetWrapper(wb_ox["Period Returns"])
                result["period_returns"] = _parse_period_returns(ws_period, datemode)
                result["annual_returns"] = _parse_annual_returns(ws_period, datemode)
            except Exception as e:
                logger.error(f"  Error parsing Period Returns: {e}")

            wb_ox.close()

        else:
            # ── xlrd path (.xls) ──
            wb = xlrd.open_workbook(filepath)
            result["file_path"] = filepath

            ws_main = wb.sheet_by_name("Composite Returns")
            as_of_val = ws_main.cell_value(1, 1)
            if as_of_val and isinstance(as_of_val, float):
                result["as_of"] = xlrd.xldate_as_datetime(as_of_val, wb.datemode)

            for strategy in ["QDVD", "SMID", "DAC", "OR"]:
                try:
                    df = _parse_composite_block(ws_main, strategy, wb.datemode)
                    if not df.empty:
                        result["composites"][strategy] = df
                        logger.info(f"  {strategy}: {len(df)} rows, "
                                    f"{df['date'].min().strftime('%Y-%m')} to "
                                    f"{df['date'].max().strftime('%Y-%m')}")
                except Exception as e:
                    logger.error(f"  Error parsing {strategy}: {e}")

            try:
                ws_period = wb.sheet_by_name("Period Returns")
                result["period_returns"] = _parse_period_returns(ws_period, wb.datemode)
                result["annual_returns"] = _parse_annual_returns(ws_period, wb.datemode)
            except Exception as e:
                logger.error(f"  Error parsing Period Returns: {e}")

        result["available"] = len(result["composites"]) > 0
        logger.info(f"Composite returns loaded: {len(result['composites'])} strategies, "
                     f"as of {result['as_of']}")

    except Exception as e:
        result["error"] = f"Error reading Composite Returns file: {e}"
        logger.error(result["error"])

    return result


# ── Convenience Functions ───────────────────────────────────────────────────

def get_monthly_returns(composite_df, return_type="gross"):
    """
    Extract a clean monthly returns series from a composite DataFrame.

    Args:
        composite_df: DataFrame from load_composite_data()["composites"][strategy]
        return_type: "gross" or "net"

    Returns:
        pd.Series indexed by date with monthly return values
    """
    col = "gross_mo" if return_type == "gross" else "net_mo"
    series = composite_df.set_index("date")[col].dropna()
    return series


def get_cumulative_series(composite_df, return_type="gross"):
    """
    Extract cumulative return series (as growth of $100).

    Returns:
        pd.Series indexed by date with cumulative value (100 = starting value)
    """
    col = "cum_gross_val" if return_type == "gross" else "cum_net_val"
    series = composite_df.set_index("date")[col].dropna()
    return series


def get_benchmark_cumulative(composite_df, which="primary"):
    """
    Extract benchmark cumulative value series.

    Args:
        which: "primary" or "secondary"

    Returns:
        tuple: (benchmark_name, pd.Series indexed by date)
    """
    if which == "primary":
        name_col, val_col = "bench1_name", "bench1_val"
    else:
        name_col, val_col = "bench2_name", "bench2_val"

    name = composite_df[name_col].iloc[0] if len(composite_df) > 0 else ""
    series = composite_df.set_index("date")[val_col].dropna()
    return name, series


def compute_risk_metrics(composite_df, return_type="gross", risk_free_rate=0.04):
    """
    Compute risk metrics from monthly composite returns.

    Args:
        composite_df: DataFrame from composites dict
        return_type: "gross" or "net"
        risk_free_rate: annual risk-free rate (default 4%)

    Returns:
        dict with: annualized_return, annualized_vol, sharpe, sortino,
                   max_drawdown, best_month, worst_month, pct_positive_months,
                   beta (vs primary benchmark)
    """
    mo_col = "gross_mo" if return_type == "gross" else "net_mo"
    monthly = composite_df[mo_col].dropna()

    if len(monthly) < 12:
        return None

    # Annualized return (geometric)
    cum = (1 + monthly).prod()
    n_years = len(monthly) / 12
    ann_return = cum ** (1 / n_years) - 1

    # Annualized volatility
    ann_vol = monthly.std() * np.sqrt(12)

    # Sharpe
    monthly_rf = (1 + risk_free_rate) ** (1/12) - 1
    excess = monthly - monthly_rf
    sharpe = (excess.mean() / excess.std()) * np.sqrt(12) if excess.std() > 0 else 0

    # Sortino (downside deviation)
    downside = excess[excess < 0]
    downside_std = np.sqrt((downside ** 2).mean()) if len(downside) > 0 else 0
    sortino = (excess.mean() * 12) / (downside_std * np.sqrt(12)) if downside_std > 0 else 0

    # Max drawdown
    cum_wealth = (1 + monthly).cumprod()
    running_max = cum_wealth.cummax()
    drawdown = (cum_wealth - running_max) / running_max
    max_dd = drawdown.min()

    # Best/worst month
    best_mo = monthly.max()
    worst_mo = monthly.min()

    # Percent positive months
    pct_positive = (monthly > 0).sum() / len(monthly)

    # Beta vs primary benchmark
    bench_mo = composite_df["bench1_mo"].dropna()
    # Align indices
    aligned = pd.concat([monthly, bench_mo], axis=1).dropna()
    if len(aligned) > 12 and aligned.iloc[:, 1].var() > 0:
        cov = aligned.iloc[:, 0].cov(aligned.iloc[:, 1])
        var = aligned.iloc[:, 1].var()
        beta = cov / var
    else:
        beta = np.nan

    # Tracking error
    if len(aligned) > 12:
        tracking_diff = aligned.iloc[:, 0] - aligned.iloc[:, 1]
        tracking_error = tracking_diff.std() * np.sqrt(12)
        info_ratio = (tracking_diff.mean() * 12) / tracking_error if tracking_error > 0 else 0
    else:
        tracking_error = np.nan
        info_ratio = np.nan

    return {
        "annualized_return": ann_return,
        "annualized_vol": ann_vol,
        "sharpe": sharpe,
        "sortino": sortino,
        "max_drawdown": max_dd,
        "best_month": best_mo,
        "worst_month": worst_mo,
        "pct_positive_months": pct_positive,
        "beta": beta,
        "tracking_error": tracking_error,
        "information_ratio": info_ratio,
    }


def build_monthly_heatmap_data(composite_df, return_type="gross"):
    """
    Build a year × month matrix for a heatmap.

    Returns:
        pd.DataFrame with years as index, month names as columns,
        values are monthly returns (as decimals, e.g., 0.05 = 5%)
    """
    mo_col = "gross_mo" if return_type == "gross" else "net_mo"
    df = composite_df[["date", mo_col]].dropna().copy()
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month

    # Pivot to year × month
    pivot = df.pivot_table(index="year", columns="month", values=mo_col, aggfunc="first")

    # Rename columns to month abbreviations
    month_names = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                   7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
    pivot = pivot.rename(columns=month_names)

    # Ensure all 12 months present
    for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]:
        if m not in pivot.columns:
            pivot[m] = np.nan

    pivot = pivot[["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]]

    # Add annual total column
    # Annual return = geometric compounding of monthly returns
    annual_data = df.groupby("year")[mo_col].apply(lambda x: (1 + x).prod() - 1)
    pivot["Annual"] = annual_data

    pivot = pivot.sort_index(ascending=False)
    return pivot